"""
The Automation Code that is performed on our webpage is within 
this file. Below are the imports that are used:
"""
import os, sys
import time
currentdir = os.path.dirname(os.path.realpath(__file__))
parentdir = os.path.dirname(currentdir)
sys.path.append(parentdir)
from locators import *
from element import BasePageElement
from selenium.webdriver.common.keys import Keys
from selenium.webdriver import ActionChains
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC, wait
from selenium.webdriver.common.by import By
from selenium.common.exceptions import NoSuchElementException 
from openpyxl.worksheet import worksheet
from openpyxl import Workbook, load_workbook
from openpyxl.styles import PatternFill, Font, Color
from openpyxl.utils import get_column_letter

class BasePage(object):
    
    def __init__(self, driver):
        self.driver = driver
        self.wait = WebDriverWait(self.driver, 30)

class MainPage(BasePage, BasePageElement):

    def is_title_matches(self):
        """returns true if title contains PEGA and false if not"""
        return 'Pega' in self.driver.title

    def login(self, username, password):
        """
        This function is used to login to the PEGA designer
        studio code the parameters are imported from main.py
        and sent to the credentials page of PEGA.
        """
        element = self.driver.find_element(*LoginPageLocators.USERNAME_TEXTBOX)
        element.send_keys(username)
        element2 = self.driver.find_element(*LoginPageLocators.PASSWORD_TEXTBOX)
        element2.send_keys(password)
        element3 = self.driver.find_element(*LoginPageLocators.LOGIN_BUTTON)
        element3.click()  

    def _unable_error(self, workbook_path):
        """
        Helper Function for our agent_toggle function. It sorts the data in the report
        and deletes rows of unwanted errors from the main error necessary for the agent_toggle
        process in the Broken Queue in Pega Designer Studio. The main error being:
        "Unable to open an instance using the given inputs: PA-Data-Location !"

        Returns: the new workbook which includes an edited version of the report/worksheet.
        """
        wb = load_workbook(workbook_path)
        ws = wb.active
        start_idx = 2
        amount = 0
        delete = False
        deleted_rows = []

        # Iterates through the report storing coordinates of rows that need to be deleted from the 
        # main error-(Unable to open an instance using the given inputs: PA-Data-Location !) required for the agent toggle
        for row in ws.iter_rows(min_row=ws.min_row+1 , max_row=ws.max_row, min_col=3, max_col=3):
            row = str(row)
            row_val = ws[(row[14:row.index('>')])].value
            if "Unable" in row_val:
                if delete == True:
                    vals = (start_idx, amount)
                    deleted_rows.append(vals)
                    delete = False
                start_idx += 1
                amount = 0
            else:
                amount += 1
                delete = True
        # Handles any error past the last Unable seeking error.
        if delete == True:
            vals = (start_idx, amount)
            deleted_rows.append(vals)
            delete = False
        start_idx += 1
        amount = 0
        # loops through the coordinates in our list-(deleted_rows) and deletes the rows
        for pair in deleted_rows:
            ws.delete_rows(idx=pair[0], amount=pair[1])

        # Handles the last row which is skipped by the first iteration through the report
        last_row = str(ws['C' + str(ws.max_row)].value)
        if "Unable" in last_row:
            pass
        else:
            ws.delete_rows(idx=ws.max_row)
        
        # Saves the new work book and returns it for later use in our requeue function
        wb.save('requeue_sheet.xlsx')
        return wb
    
    def _agent_toggle_list(self, workbook_path):
        """
        Helper funciton imports an excel spreadsheet that has been filtered
        in our _unable_error helper method above. It loops through
        the data extracting policy numbers and appends it to our policy_num_list.

        Returns: A list of policy numbers to be used in our agent_toggle function
        """
        new_wb = self._unable_error(workbook_path)
        ws = new_wb.active
        policy_num_list = []

        #iterates over the xcel sheet produced by the _unable_error function and 
        #stores the policy numbers in policy_num_list
        for row in ws.iter_rows(min_row=ws.min_row+1 , max_row=ws.max_row, min_col=4, max_col=4):
            row = str(row)
            policy = ws[row[14:row.index('>')]].value
            policy_line = policy.split()
            policy_num_list.append(policy_line[1])

        print(policy_num_list)
        return(policy_num_list)

    def _admin_portal(self):
        """
        Helper funtion for our agent toggle. The agent toggle is a process
        that takes place in the Residential Admin Portal which is accessed through
        PEGA designer studio. This function opens the Residential admin portal for 
        the agent_toggle funciton.
        """
        self.driver.implicitly_wait(10)
        self.wait.until(EC.presence_of_element_located((By.XPATH, '/html/body/div[4]/header/div/div/div[1]/div/div/div[3]'))).click()
        
        self.driver.implicitly_wait(10)
        switch_app = self.driver.find_element(*DesinerStudioLocators.SWITCH_APP)
        residential = self.driver.find_element(*DesinerStudioLocators.RESIDENTIAL)
        actions = ActionChains(self.driver)
        actions.move_to_element(switch_app).move_to_element(residential).click().perform()
        time.sleep(3)

        launch_button = self.driver.find_element(*DesinerStudioLocators.LAUNCH_BUTTON)
        launch_button.click()

        admin_portal_button = self.driver.find_element(*DesinerStudioLocators.ADMIN_PORTAL)
        actions = ActionChains(self.driver)
        actions.move_to_element(admin_portal_button).click().perform()
        self.driver.get('https://palomr-psic-prod1.pegacloud.net/prweb/c_zdp4FjGamBkvISSFNU2z_tkf4x_vT1f179hYUjuHw%5B*/!STANDARD?pyActivity=Embed-PortalLayout.RedirectAndRun&ThreadName=OpenPortal&Location=pyActivity%3DData-Portal.ShowSelectedPortal%26portal%3DPARLAdminPortal%26Name%3D%20PARLAdminPortal%26pzSkinName%3D%26developer%3Dfalse%26ThreadName%3DOpenPortal%26launchPortal%3Dtrue&bPurgeTargetThread=true&target=popup&pzHarnessID=HIDABE5A048FEF4AD9D617954DF4C5BE656')
    def _check_path(self, path, locator):
        """
        This helper function will check if our xpaths exist and will return True or False.
        """
        try:
            self.driver.find_element(path, locator)
        except NoSuchElementException:
            return False
        return True

    def agent_toggle(self, workbook_path):
        """
        This funciton imports an excel spreadsheet and loops through
        the data extracting policy numbers and then it automates the agent
        toggle proccess used on the front end of residential admin portal in PEGA.

        Parameters: (workbook_path: type-(excel file path)) file that contains the data
        used to manipulate and perform our agent toggle process.
        """
        #calls to our helper functions in order to filter the data
        self._admin_portal()
        policy_num_list = self._agent_toggle_list(workbook_path)

        #lists that store policies based on the condition of the policies
        renewed_list = [] # contains a list of policies that have been renewed
        issues_list = []  # contains a list of policies that have issues and need to be reviewed manually
        requeue_list = [] # contains a list of policies that have gone through the agent toggle process successfully

        #iterates over our policy_num_list performing the agent toggle process 
        for policy in policy_num_list:
            # 1) Enter Policy Number into Search Box
            search_box = self.driver.find_element(*AdminPortalLocators.SEARCH_POLICY_BOX)
            search_box.clear()
            time.sleep(.5)
            search_box.send_keys(policy)
            actions = ActionChains(self.driver)
            actions.move_to_element(search_box).send_keys(Keys.RETURN).perform()
            
            # switched frame to access action drop down menu and other elements located in frame below.
            time.sleep(.5)
            self.driver.switch_to.frame("PegaGadget1Ifr")
            time.sleep(2)
            self.driver.find_element(*AdminPortalLocators.ACTIONS_DROP_MENU).click()
            
            # 2)Checks if the policy has been renewed already and closes if True.
            if (self._check_path(*AdminPortalLocators.RENEWED) and self.driver.find_element(*AdminPortalLocators.RENEWED).is_displayed())\
                or self._check_path(*AdminPortalLocators.RENEWED_CHECK_BOX):
                    print(f'Policy: {policy} has been renewed already!')
                    renewed_list.append(policy)
                    self.driver.find_element(*AdminPortalLocators.CLOSE).click()
            elif (self._check_path(*AdminPortalLocators.RENEWED_2) and self.driver.find_element(*AdminPortalLocators.RENEWED_2).is_displayed())\
                or self._check_path(*AdminPortalLocators.RENEWED_CHECK_BOX):
                    print(f'Policy: {policy} has been renewed already!')
                    renewed_list.append(policy)
                    self.driver.find_element(*AdminPortalLocators.CLOSE).click()
            
            # 3)If update shell button is missing the policy will be closed.
            elif self._check_path(*AdminPortalLocators.UPDATE_SHELL) == False:
                print(f'Policy: {policy} Action Drop Menu Issue. Update Shell button missing and policy was not renewed.')
                issues_list.append(policy)
                self.driver.find_element(*AdminPortalLocators.CLOSE).click() 

            # 4) Policy meets all criteria for renewal and will go through the agent toggle process in the following code.
            else:
                ho_df_box_check = self.driver.find_element(*AdminPortalLocators.HO_DF_BOX_CHECK)
                home_policy_num_check = self.driver.find_element(*AdminPortalLocators.HOME_POLICY_NUM_CHECK)
                agent_check = self.driver.find_element(*AdminPortalLocators.AGENT_CHECK).text
                print(ho_df_box_check.text)
                print(home_policy_num_check.text)
                print(agent_check)
                ho_df_text = ho_df_box_check.text
                home_policy_text = home_policy_num_check.text

                # Policy gets checked for empty fields and is closed if True or continue through the agent toggle if False.    
                if (ho_df_text == '' and home_policy_text != 'Not Available') or (ho_df_text == '' and home_policy_text == '') or (agent_check == ''):
                    print(f'Policy: {policy} has Empty HO/DF Carrier or Missing Fields and the policy was not renewed.')
                    issues_list.append(policy)
                elif self._check_path(*AdminPortalLocators.UPDATE_SHELL):
                    self.driver.find_element(*AdminPortalLocators.UPDATE_SHELL).click()
                    time.sleep(1)
                    ho_df_box = self.driver.find_element(*AdminPortalLocators.HO_DF_BOX)
                    if ho_df_text == '' and home_policy_text == 'Not Available': # Special case were HO/DF box needs input-"Other".
                        ho_df_box.send_keys("Other")
                        other_button = self.driver.find_element(*AdminPortalLocators.OTHER_BUTTON)
                        self.driver.execute_script("arguments[0].click();", other_button)
                    self.driver.implicitly_wait(10)
                    token = self.driver.find_element(*AdminPortalLocators.AGENT_CODE).text
                    tokens = token.split('-')
                    agent_code = f'{tokens[0]}-{tokens[1]}-{tokens[2]}'
                    if agent_code == "PSIC-658-28": #Special Case were agent code PSIC-658-1 is used because previous code does not exist anymore.
                        agent_code = "PSIC-658-1"
                    print(agent_code, policy) #For user verification purposes DO NOT DELETE
                    agent_name = self.driver.find_element(*AdminPortalLocators.AGENCY_NAME)
                    agent_name.send_keys('1341')
                    agent_name.send_keys(Keys.SPACE)
                    self.wait.until(EC.element_to_be_clickable((By.XPATH, '//tr[contains(@data-gargs, "John MacDonald")]'))).click()
                    self.wait.until(EC.element_to_be_clickable((By.ID, 'ContactId'))).click()
                    self.driver.find_element_by_xpath('//option[contains(@value, "PSIC-1341-1")]').click()
                    self.driver.find_element(*AdminPortalLocators.SUBMIT).click()
                    time.sleep(1)
                    self.driver.find_element(*AdminPortalLocators.ACTIONS_DROP_MENU).click()
                    time.sleep(2)
                    self.driver.find_element(*AdminPortalLocators.UPDATE_SHELL).click()
                    time.sleep(2)
                    agent_name = self.driver.find_element(*AdminPortalLocators.AGENCY_NAME)
                    agent_name.send_keys(agent_code)
                    agent_name.send_keys(Keys.SPACE)
                    self.driver.find_element_by_xpath(f'//tr[contains(@data-gargs, "{agent_code}")]').click()
                    time.sleep(2)
                    self.driver.find_element(*AdminPortalLocators.AGENT_ON_RECORD).click()
                    self.driver.find_element_by_xpath(f'//option[contains(@value, "{agent_code}")]').click()
                    self.driver.find_element(*AdminPortalLocators.SUBMIT).click()
                    time.sleep(2)
                    print(f'Policy: {policy} has gone through the agent toggle process and needs to be requeued')
                    requeue_list.append(policy)
                self.driver.find_element(*AdminPortalLocators.ACTIONS_DROP_MENU).click()
                time.sleep(2)
                self.driver.find_element(*AdminPortalLocators.CLOSE).click()
                

        wb = self._unable_error(workbook_path)
        ws = wb.active
        
        #iterates over the xcel file marking rows based on three conditions
        #1)Green-The policy has been renewed already
        #2)Yellow-The policy has issues and needs to be checked manually
        #3)White-The policy has gone through the agent toggle successfully and needs to be requeued
        for row in ws.iter_rows(min_row=ws.min_row+1 , max_row=ws.max_row, min_col=0, max_col=4):
            val = str(row)
            token = val.split(',')
            row_val = ws[val[14:val.index('>')]]
            policy_val = (token[3][14:token[3].index(">")])
            policy_line = ws[policy_val].value
            policy = policy_line.split()
            if policy[1] in renewed_list:
                row_val.fill = PatternFill(start_color='FF92D050', 
                                        end_color='FF92D050',
                                        fill_type='solid')
            elif policy[1] in issues_list:
                row_val.fill = PatternFill(start_color="FFFFFF00",
                                        end_color="FFFFFF00",
                                        fill_type='solid')
        wb.save('requeue_sheet.xlsx')
        # prints out the lists of policies based on 
        # the condition of the policy and is to help 
        # in terms of verification purposes.
        print(issues_list)
        print(renewed_list)
        print(requeue_list)
        # wb.save('requeue_sheet.xlsx')

class StudioPage(BasePage, BasePageElement):
    
    def login(self, username, password):
        """
        This function automates the login process into PEGA.
        """
        element = self.driver.find_element(*LoginPageLocators.USERNAME_TEXTBOX)
        element.send_keys(username)
        element2 = self.driver.find_element(*LoginPageLocators.PASSWORD_TEXTBOX)
        element2.send_keys(password)
        element3 = self.driver.find_element(*LoginPageLocators.LOGIN_BUTTON)
        element3.click()
        
    def requeue(self, requeue_worksheet):
        """
        This function will take all of the policies that 
        have gone through the agent toggle process and requeue
        them in the Queue Management. This is found in Residential 
        Policies in Pega Designer Studio Code. 
        """
        # Variables for our XCEL Sheet, RGB values to identify rows, and requeue list
        wb = load_workbook(requeue_worksheet)
        ws = wb.active
        yellow = 'FFFFFF00'
        green = 'FF92D050'
        requeue_lst = []

        # Loops through the XCEL report grabbing the system_id number based on the color of the row.
        # If the row is white it will append the system_id from column 1 to our requeue list.
        for row in ws.iter_rows(min_row=ws.min_row+1 , max_row=ws.max_row, min_col=1, max_col=1):
            row = str(row)
            row_color_rgb = ws[(row[14:row.index('>')])].fill.start_color.rgb
            if row_color_rgb != yellow and row_color_rgb != green:
                system_id = str(ws[row[14:row.index('>')]].value)
                requeue_lst.append(system_id)
        
        # Xpaths for our buttons
        default_entry_button = '/html/body/div[2]/form/div[3]/div/table/tbody/tr/td/div/div[2]/div/div/div/div[11]/div[2]/div/div/div/div/span/div/div/div/div/div[1]/div/div/div/div/div/div/div[2]/div/div/div/div/div/div[1]/div/table/tbody/tr/td[2]/div/table/tbody/tr[3]'
        filter_btn = '/html/body/div[2]/form/div[3]/div/table/tbody/tr/td/div/div[2]/div/div/div/div[11]/div[2]/div/div/div/div/span/div/div/div/div/div[2]/div/div/div/div/div/div/div/div/div[2]/div/div/div/div/div/div/div[1]/table/tbody/tr/td[2]/div/table/tbody/tr/th[2]/div/span/a'
        search_box = '/html/body/div[2]/form/div[5]/div[1]/ul/li[1]/div[2]/div/span/div/div/div/div/div/div/div/div/div/div/div/div/div/span/input'
        apply_btn = '/html/body/div[2]/form/div[5]/div[1]/ul/li[3]/div/button[1]'
        checkbox = '/html/body/div[2]/form/div[3]/div/table/tbody/tr/td/div/div[2]/div/div/div/div[11]/div[2]/div/div/div/div/span/div/div/div/div/div[2]/div/div/div/div/div/div/div/div/div[2]/div/div/div/div/div/div/div/div[1]/div/table/tbody/tr/td[2]/div/table/tbody/tr[2]/td[1]/div/input[2]'
        requeue_btn = '/html/body/div[2]/form/div[3]/div/table/tbody/tr/td/div/div[2]/div/div/div/div[11]/div[2]/div/div/div/div/span/div/div/div/div/div[2]/div/div/div/div/div/div/div/div/div[1]/div/div/div/div/div[2]/div/div/div/div[1]/div/div/div/div[1]/span/button'
        self.driver.get('https://palomr-psic-prod1.pegacloud.net/prweb/c_zdp4FjGaneiE7ov3NHrVUn18ZVkv4_MIcwQEnLUJ4%5B*/!STANDARD?pyActivity=Data-Portal.ShowDesktop')

        self.driver.implicitly_wait(10)
        designer_studio = self.driver.find_element_by_xpath('//*[@id="RULE_KEY"]/div/div/div/div/div/div/span/a')
        designer_studio.click()

        actions = ActionChains(self.driver)
        systems = self.driver.find_element_by_xpath('//*[@id="menu-item-$PpyNavigation1623382803785$ppyElements$l1$ppyElements$l10"]/a')
        operations = self.driver.find_element_by_xpath('//*[@id="menu-item-$PpyNavigation1623382803785$ppyElements$l1$ppyElements$l10$ppyElements$l8"]')
        queue_management = self.driver.find_element_by_xpath('//*[@id="$PpyNavigation1623382803785$ppyElements$l1$ppyElements$l10$ppyElements$l8"]/li[5]/a')
        actions.move_to_element(systems).move_to_element(operations).move_to_element(queue_management).click().perform()
        
        wait_time = WebDriverWait(self.driver, 30)
        
        # Loops through our requeue list and requeues
        # policies that have gone through the agent toggle process successfully.
        if requeue_lst !=0:
            for system in requeue_lst:
                wait_time.until(EC.frame_to_be_available_and_switch_to_it('PegaGadget0Ifr'))
                wait_time.until(EC.presence_of_element_located((By.XPATH, default_entry_button))).click()
                time.sleep(2)
                self.driver.find_element_by_xpath(filter_btn).click()
                time.sleep(2)
                print(system)
                wait_time.until(EC.presence_of_element_located((By.XPATH, search_box))).send_keys(system)
                self.driver.find_element_by_xpath(apply_btn).click()
                self.driver.implicitly_wait(10)
                self.driver.find_element_by_xpath(checkbox).click()
                time.sleep(2)
                self.driver.find_element_by_xpath(requeue_btn).click()
                self.driver.refresh()
        else:
            print('All policies have been requeued')

    def _goal_seeking_error(self, workbook_path):
        """
        Helper Function for our removal function. It sorts the data in the report
        and deletes rows of unwanted errors from the main error necessary for the removal
        process in the Broken Queue in Pega Designer Studio. The main error being:
        "Goal seek requires missing input property RewnewedPolicyCase.IsRenewalSuccessful on page RewnewedPolicyCase; details: (unknown)"

        Returns: A new workbook which includes an edited version of the report/worksheet.
        """
        wb = load_workbook(workbook_path)
        ws = wb.active
        start_idx = 2
        amount = 0
        delete = False
        deleted_rows = []

        # Iterates through the report storing coordinates of rows that need to be deleted from the 
        # the goal seeking error necessary for the removal process.
        for row in ws.iter_rows(min_row=ws.min_row+1 , max_row=ws.max_row, min_col=3, max_col=3):
            row = str(row)
            row_val = ws[(row[14:row.index('>')])].value
            if "Goal" in row_val:
                if delete == True:
                    vals = (start_idx, amount)
                    deleted_rows.append(vals)
                    delete = False
                start_idx += 1
                amount = 0
            else:
                amount += 1
                delete = True
        # Handles any error past the last goal seeking error.
        if delete == True:
            vals = (start_idx, amount)
            deleted_rows.append(vals)
            delete = False
        start_idx += 1
        amount = 0
        # loops through the coordinates in our list-(deleted_rows) and deletes the rows
        for pair in deleted_rows:
            ws.delete_rows(idx=pair[0], amount=pair[1])

        # Handles the last row which is skipped by the first iteration through the report
        last_row = str(ws['C' + str(ws.max_row)].value)
        if "Goal" in last_row:
            pass
        else:
            ws.delete_rows(idx=ws.max_row)
        
        # Saves the new work book and returns it for later use in our removal function
        wb.save('removal_sheet.xlsx')
        return wb

    def remove(self, removal_worksheet):
        """
        This function will take all of the goal seeking errors and remove
        them from the Broken Queue. This is located in Residential 
        Policies in Pega Designer Studio Code. 
        """
        # Variables for our XCEL Sheet, RGB values to identify rows, and requeue list
        wb = self._goal_seeking_error(removal_worksheet)
        ws = wb.active
        remove_lst = []

        # Loops through the XCEL report grabbing the system_id number based on the color of the row.
        # If the row is white it will append the system_id from column 1 to our requeue list.
        for row in ws.iter_rows(min_row=ws.min_row+1 , max_row=ws.max_row, min_col=1, max_col=1):
            row = str(row)
            system_id = str(ws[row[14:row.index('>')]].value)
            remove_lst.append(system_id)
        
        # Xpaths for our buttons
        default_entry_button = '/html/body/div[2]/form/div[3]/div/table/tbody/tr/td/div/div[2]/div/div/div/div[11]/div[2]/div/div/div/div/span/div/div/div/div/div[1]/div/div/div/div/div/div/div[2]/div/div/div/div/div/div[1]/div/table/tbody/tr/td[2]/div/table/tbody/tr[3]'
        filter_btn = '/html/body/div[2]/form/div[3]/div/table/tbody/tr/td/div/div[2]/div/div/div/div[11]/div[2]/div/div/div/div/span/div/div/div/div/div[2]/div/div/div/div/div/div/div/div/div[2]/div/div/div/div/div/div/div[1]/table/tbody/tr/td[2]/div/table/tbody/tr/th[2]/div/span/a'
        search_box = '/html/body/div[2]/form/div[5]/div[1]/ul/li[1]/div[2]/div/span/div/div/div/div/div/div/div/div/div/div/div/div/div/span/input'
        apply_btn = '/html/body/div[2]/form/div[5]/div[1]/ul/li[3]/div/button[1]'
        checkbox = '/html/body/div[2]/form/div[3]/div/table/tbody/tr/td/div/div[2]/div/div/div/div[11]/div[2]/div/div/div/div/span/div/div/div/div/div[2]/div/div/div/div/div/div/div/div/div[2]/div/div/div/div/div/div/div/div[1]/div/table/tbody/tr/td[2]/div/table/tbody/tr[2]/td[1]/div/input[2]'
        remove_btn = '/html/body/div[2]/form/div[3]/div/table/tbody/tr/td/div/div[2]/div/div/div/div[11]/div[2]/div/div/div/div/span/div/div/div/div/div[2]/div/div/div/div/div/div/div/div/div[1]/div/div/div/div/div[2]/div/div/div/div[2]/div/div/div/div[1]/span/button'
        self.driver.get('https://palomr-psic-prod1.pegacloud.net/prweb/c_zdp4FjGaneiE7ov3NHrVUn18ZVkv4_MIcwQEnLUJ4%5B*/!STANDARD?pyActivity=Data-Portal.ShowDesktop')

        self.driver.implicitly_wait(10)
        designer_studio = self.driver.find_element_by_xpath('//*[@id="RULE_KEY"]/div/div/div/div/div/div/span/a')
        designer_studio.click()

        actions = ActionChains(self.driver)
        systems = self.driver.find_element_by_xpath('//*[@id="menu-item-$PpyNavigation1623382803785$ppyElements$l1$ppyElements$l10"]/a')
        operations = self.driver.find_element_by_xpath('//*[@id="menu-item-$PpyNavigation1623382803785$ppyElements$l1$ppyElements$l10$ppyElements$l8"]')
        queue_management = self.driver.find_element_by_xpath('//*[@id="$PpyNavigation1623382803785$ppyElements$l1$ppyElements$l10$ppyElements$l8"]/li[5]/a')
        actions.move_to_element(systems).move_to_element(operations).move_to_element(queue_management).click().perform()
        
        wait_time = WebDriverWait(self.driver, 30)
        
        # Loops through our requeue list and requeues
        # policies that have gone through the agent toggle process successfully.
        if remove_lst !=0:
            for system_id in remove_lst:
                wait_time.until(EC.frame_to_be_available_and_switch_to_it('PegaGadget0Ifr'))
                wait_time.until(EC.presence_of_element_located((By.XPATH, default_entry_button))).click()
                time.sleep(2)
                self.driver.find_element_by_xpath(filter_btn).click()
                time.sleep(2)
                print(system_id, "removed")
                wait_time.until(EC.presence_of_element_located((By.XPATH, search_box))).send_keys(system_id)
                self.driver.find_element_by_xpath(apply_btn).click()
                self.driver.implicitly_wait(10)
                self.driver.find_element_by_xpath(checkbox).click()
                time.sleep(2)
                self.driver.find_element_by_xpath(remove_btn).click()
                self.driver.refresh()